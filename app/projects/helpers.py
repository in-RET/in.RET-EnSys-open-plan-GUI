import json
import os
import io
import csv
from openpyxl import load_workbook
from django import forms
from django.contrib.staticfiles.storage import staticfiles_storage
from django.core.exceptions import ValidationError
from django.utils.translation import ugettext_lazy as _
from projects.dtos import convert_to_dto
from projects.constants import MAP_MVS_EPA
from dashboard.helpers import KPIFinder

from oemof.tools import economics

PARAMETERS = {}
if os.path.exists(staticfiles_storage.path("MVS_parameters_list.csv")) is True:
    with open(
        staticfiles_storage.path("MVS_parameters_list.csv"), encoding="utf-8"
    ) as csvfile:
        csvreader = csv.reader(csvfile, delimiter=",", quotechar='"')
        for i, row in enumerate(csvreader):
            if i == 0:
                hdr = row
                label_idx = hdr.index("label")
            else:
                label = row[label_idx]
                label = MAP_MVS_EPA.get(label, label)
                PARAMETERS[label] = {}
                for k, v in zip(hdr, row):
                    if k == "sensitivity_analysis":
                        v = bool(int(v))
                    PARAMETERS[label][k] = v

parameters_helper = KPIFinder(param_info_dict=PARAMETERS, unit_header=":Unit:")


# Helper method to clean dict data from empty values
def remove_empty_elements(d):
    def empty(x):
        return x is None or x == {} or x == []

    if not isinstance(d, (dict, list)):
        return d
    elif isinstance(d, list):
        return [v for v in (remove_empty_elements(v) for v in d) if not empty(v)]
    else:
        return {
            k: v
            for k, v in ((k, remove_empty_elements(v)) for k, v in d.items())
            if not empty(v)
        }


# Helper to convert Scenario data to MVS importable json
def format_scenario_for_mvs(scenario_to_convert, testing=False):
    mvs_request_dto = convert_to_dto(scenario_to_convert, testing=testing)
    print(mvs_request_dto)
    dumped_data = json.loads(
        json.dumps(mvs_request_dto.__dict__, default=lambda o: o.__dict__)
    )

    # format the constraints in MVS format directly, thus avoiding the need to maintain MVS-EPA
    # parser in multi-vector-simulator package
    constraint_dict = {}
    for constraint in dumped_data["constraints"]:
        constraint_dict[constraint["label"]] = constraint["value"]
    dumped_data["constraints"] = constraint_dict

    # Remove None values
    return dumped_data  # remove_empty_elements


Zinssatz = 10.5


def epc_calc(capex, Amortisierungszeit, opex):
    investk = economics.annuity(capex=capex, n=Amortisierungszeit, wacc=Zinssatz / 100)
    betriebsk = capex * (opex / 100)
    epc = investk + betriebsk
    return epc


def sensitivity_analysis_payload(
    variable_parameter_name="",
    variable_parameter_range="",
    variable_parameter_ref_val="",
    output_parameter_names=None,
):
    """format the parameters required to request a sensitivity analysis in a specific JSON"""
    if output_parameter_names is None:
        output_parameter_names = []
    return {
        "sensitivity_analysis_settings": {
            "variable_parameter_name": variable_parameter_name,
            "variable_parameter_range": variable_parameter_range,
            "variable_parameter_ref_val": variable_parameter_ref_val,
            "output_parameter_names": output_parameter_names,
        }
    }


SA_RESPONSE_SCHEMA = {
    "type": "object",
    "required": ["server_info", "mvs_version", "id", "status", "results"],
    "properties": {
        "server_info": {"type": "string"},
        "mvs_version": {"type": "string"},
        "id": {"type": "string"},
        "status": {"type": "string"},
        "results": {
            "type": "object",
            "required": ["reference_simulation_id", "sensitivity_analysis_steps"],
            "properties": {
                "reference_simulation_id": {"type": "string"},
                "sensitivity_analysis_steps": {
                    "type": "array",
                    "items": {"type": "object"},
                },
            },
            "additionalProperties": False,
        },
        "ref_sim_id": {"type": "string"},
        "sensitivity_analysis_ids": {"type": "array", "items": {"type": "string"}},
    },
    "additionalProperties": False,
}


# Used to proof the json objects stored as text in the db
SA_OUPUT_NAMES_SCHEMA = {"type": "array", "items": {"type": "string"}}


def sa_output_values_schema_generator(output_names):
    return {
        "type": "object",
        "required": output_names,
        "properties": {
            output_name: {
                "type": "object",
                "required": ["value", "path"],
                "properties": {
                    "value": {
                        "oneOf": [
                            {"type": "null"},
                            {
                                "type": "array",
                                "items": {
                                    "anyOf": [{"type": "number"}, {"type": "null"}]
                                },
                            },
                        ]
                    },
                    "path": {
                        "oneOf": [
                            {"type": "string"},
                            {"type": "array", "items": {"type": "string"}},
                        ]
                    },
                },
            }
            for output_name in output_names
        },
        "additionalProperties": False,
    }


class DualInputWidget(forms.MultiWidget):

    template_name = "asset/dual_input.html"

    def __init__(self, **kwargs):
        """This special input consist of one text field and one upload file button"""

        self.default = kwargs.pop("default", None)
        self.param_name = kwargs.pop("param_name", None)

        widgets = {
            "scalar": forms.TextInput(
                attrs={
                    "class": "form-control",
                    "onchange": f"plotDualInputTrace(obj=this.value, param_name='{self.param_name}')",
                }
            ),
            "file": forms.FileInput(
                attrs={
                    "class": "form-control",
                    "onchange": f"uploadDualInputTrace(obj=this.files, param_name='{self.param_name}')",
                }
            ),
        }
        super(DualInputWidget, self).__init__(widgets=widgets, **kwargs)

    def use_required_attribute(self, initial):
        # overwrite the method of the Widget class of the django.form.widgets module
        return False

    def decompress(self, value):
        answer = [self.default, None]
        if value is not None:
            value = json.loads(value)
            answer = [value, None]
        return answer


class DualNumberField(forms.MultiValueField):
    def __init__(self, default=None, param_name=None, **kwargs):
        fields = (forms.DecimalField(required=False), forms.CharField(required=False))
        kwargs.pop("max_length", None)
        self.min = kwargs.pop("min", None)
        self.max = kwargs.pop("max", None)
        kwargs["widget"] = DualInputWidget(default=default, param_name=param_name)

        super().__init__(fields=fields, require_all_fields=False, **kwargs)

    def clean(self, values):
        """If a file is provided it will be considered over the scalar"""
        scalar_value, timeseries_file = values

        if timeseries_file is not None:
            input_timeseries_values = parse_input_timeseries(timeseries_file)
            answer = input_timeseries_values
        else:

            # check the input string is a number or a list
            if scalar_value != "":
                try:
                    answer = float(scalar_value)
                except ValueError:
                    try:
                        answer = json.loads(scalar_value)
                        if not isinstance(answer, list):
                            scalar_value = ""
                    except json.decoder.JSONDecodeError:
                        scalar_value = ""

            if scalar_value == "":
                self.set_widget_error()
                raise ValidationError(
                    _(
                        "Please provide either a number within %(boundaries) s or upload a timeseries from a file"
                    ),
                    code="required",
                    params={"boundaries": self.boundaries},
                )
        self.check_boundaries(answer)
        return answer

    @property
    def boundaries(self):
        if self.min is not None:
            min_val = self.min
        else:
            min_val = "-inf"

        if self.max is not None:
            max_val = self.max
        else:
            max_val = "inf"

        return f"[{min_val}, {max_val}]"

    def check_boundaries(self, value):

        boundaries = self.boundaries
        if isinstance(value, list):
            for v in value:
                try:
                    self.check_boundaries(v)
                except ValidationError:
                    self.set_widget_error()
                    raise ValidationError(
                        _(
                            "Some values in the timeseries do not lie within %(boundaries) s, please check your input again."
                        ),
                        code="invalid",
                        params={"boundaries": boundaries},
                    )

        else:
            if self.min is not None:
                if value < self.min:
                    self.set_widget_error()
                    raise ValidationError(
                        _("Please enter a value within %(boundaries) s"),
                        code="invalid",
                        params={"boundaries": boundaries},
                    )

            if self.max is not None:
                if value > self.max:
                    self.set_widget_error()
                    raise ValidationError(
                        _("Please enter a value within %(boundaries) s"),
                        code="invalid",
                        params={"boundaries": boundaries},
                    )

    def set_widget_error(self):
        for widget in self.widget.widgets:
            css = widget.attrs.get("class", None)
            if css is not None:
                css = css.split(" ")
            else:
                css = []
            css.append("is-invalid")
            widget.attrs["class"] = " ".join(css)


def parse_csv_timeseries(file_str):
    io_string = io.StringIO(file_str)
    delimiter = ","
    if file_str.count(";") > 0:
        delimiter = ";"

    # check if the number of , is an integer time the number of line return
    # if not, the , is probably not a column separator and a decimal separator indeed
    if file_str.count(",") % (file_str.count("\n") + 1) != 0:
        delimiter = ";"

    reader = csv.reader(io_string, delimiter=delimiter)
    timeseries_values = []
    for row in reader:
        if len(row) == 1:
            value = row[0]
        else:
            # assumes the first row is timestamps and read the second one, ignore any other row
            value = row[1]
        # convert potential comma used as decimal point to decimal point
        timeseries_values.append(float(value.replace(",", ".")))
    return timeseries_values


def parse_input_timeseries(timeseries_file):
    if timeseries_file.name.endswith("xls") or timeseries_file.name.endswith("xlsx"):
        wb = load_workbook(filename=timeseries_file)
        worksheet = wb.active
        timeseries_values = []
        n_col = worksheet.max_column

        col_idx = 0

        if n_col > 1:
            col_idx = 1

        for j in range(0, worksheet.max_row):
            try:
                timeseries_values.append(
                    float(worksheet.cell(row=j + 1, column=col_idx + 1).value)
                )
            except ValueError:
                pass

    else:
        timeseries_file_str = timeseries_file.read().decode("utf-8")

        if timeseries_file_str != "":
            if timeseries_file.name.endswith("json"):
                timeseries_values = json.loads(timeseries_file_str)
            elif timeseries_file.name.endswith("csv"):
                timeseries_values = parse_csv_timeseries(timeseries_file_str)

            elif timeseries_file.name.endswith("txt"):
                nlines = timeseries_file_str.count("\n") + 1
                if nlines == 1:
                    timeseries_values = json.loads(timeseries_file_str)
                else:
                    timeseries_values = parse_csv_timeseries(timeseries_file_str)
            else:
                raise TypeError(
                    _(
                        f'Input timeseries file type of "{timeseries_file.name}" is not supported. The supported formats are "json", "csv", "txt", "xls" and "xlsx"'
                    )
                )
        else:
            raise ValidationError(
                _('Input timeseries file "%(fname)s" is empty'),
                code="empty_file",
                params={"fname": timeseries_file.name},
            )
    return timeseries_values
